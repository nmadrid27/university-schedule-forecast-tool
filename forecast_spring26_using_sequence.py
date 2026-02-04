#!/usr/bin/env python3

import argparse
import csv
import difflib
import math
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook


FOUN_CODE_RE = re.compile(r"\bFOUN\s*(\d{3})\b", re.IGNORECASE)
SHEET_RE = re.compile(r"^(?P<term>\d{6})\s*-\s*(?P<campus>[A-Z]{3})\s*-\s*(?P<stype>FR|TR)\s*$")


def normalize_text(value: str) -> str:
    value = value or ""
    value = str(value).upper()
    value = value.replace("&", " AND ")
    value = value.replace("/", " ")
    value = value.replace("-", " ")
    value = re.sub(r"[()]", " ", value)
    value = re.sub(r"^\d+\s+", "", value)  # e.g. "1 Acting"
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def normalize_program_base(program: str) -> str:
    program = normalize_text(program)
    program = program.replace("ON LY", "ONLY")  # typo observed in source
    program = re.sub(r"\b(FALL|WINTER|SPRING)\s+ONLY\b", " ", program)
    program = re.sub(r"\bGROUP\s+[A-Z]\b", " ", program)
    program = re.sub(r"\s+", " ", program).strip()
    return program


def extract_foun_codes(cell_value: str) -> list[str]:
    if not cell_value:
        return []
    codes = []
    for match in FOUN_CODE_RE.findall(str(cell_value)):
        codes.append(f"FOUN {match}")
    # Deduplicate while preserving order
    seen = set()
    deduped = []
    for code in codes:
        if code in seen:
            continue
        seen.add(code)
        deduped.append(code)
    return deduped


@dataclass(frozen=True)
class QuarterPlan:
    mode: str  # "required" | "choice"
    courses: tuple[str, ...]


@dataclass(frozen=True)
class SequenceRow:
    program_raw: str
    program_base: str
    degree: str
    campus_raw: str
    campuses: tuple[str, ...]
    year: str
    fall: QuarterPlan
    winter: QuarterPlan
    spring: QuarterPlan
    summer: QuarterPlan


def parse_quarter_plan(cell_value: str) -> QuarterPlan:
    text = str(cell_value).strip() if cell_value else ""
    if not text:
        return QuarterPlan(mode="required", courses=tuple())
    mode = "choice" if "CHOICE" in text.upper() else "required"
    courses = tuple(extract_foun_codes(text))
    return QuarterPlan(mode=mode, courses=courses)


def parse_campuses(campus_raw: str) -> tuple[str, ...]:
    campus_norm = normalize_text(campus_raw)
    if not campus_norm:
        return tuple()
    if campus_norm == "GENERAL":
        return ("GENERAL",)
    parts = [p.strip() for p in campus_norm.split("|")]
    parts = [p for p in parts if p]
    return tuple(parts)


def campus_matches(sequence_row: SequenceRow, campus: str) -> bool:
    campus = normalize_text(campus)
    if not campus:
        return False
    if "GENERAL" in sequence_row.campuses:
        return True
    return campus in sequence_row.campuses


def load_sequence_rows(path: Path) -> list[SequenceRow]:
    rows: list[SequenceRow] = []
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for r in reader:
            campus_raw = (r.get("campus") or "").strip()
            program_raw = (r.get("program") or "").strip()
            degree = (r.get("degree") or "").strip()
            year = (r.get("year") or "").strip()
            if not program_raw or not year:
                continue
            rows.append(
                SequenceRow(
                    program_raw=program_raw,
                    program_base=normalize_program_base(program_raw),
                    degree=degree,
                    campus_raw=campus_raw,
                    campuses=parse_campuses(campus_raw),
                    year=year,
                    fall=parse_quarter_plan(r.get("fall")),
                    winter=parse_quarter_plan(r.get("winter")),
                    spring=parse_quarter_plan(r.get("spring")),
                    summer=parse_quarter_plan(r.get("summer")),
                )
            )
    return rows


def build_program_index(sequence_rows: Iterable[SequenceRow]) -> dict[str, list[SequenceRow]]:
    index: dict[str, list[SequenceRow]] = defaultdict(list)
    for row in sequence_rows:
        index[row.program_base].append(row)
    return dict(index)


def best_program_matches(program_index: dict[str, list[SequenceRow]], program_name: str) -> list[str]:
    name = normalize_program_base(program_name)
    if not name:
        return []
    if name in program_index:
        return [name]

    candidates = []
    for key in program_index.keys():
        if key in name or name in key:
            candidates.append(key)
    if candidates:
        # Prefer the shortest containing key (more specific than very long keys)
        candidates.sort(key=len)
        return candidates

    # Fuzzy fallback
    keys = list(program_index.keys())
    close = difflib.get_close_matches(name, keys, n=5, cutoff=0.78)
    return close


def parse_admissions_counts(
    admissions_path: Path,
    terms: set[str],
) -> dict[tuple[str, str, str], dict[str, dict[str, int]]]:
    """
    Returns:
      counts[(term, campus_name, stype)][scenario][interest_name] = count
    where scenario in {"accepted", "confirmed"}.
    """
    wb = load_workbook(admissions_path, read_only=True, data_only=True)
    counts: dict[tuple[str, str, str], dict[str, dict[str, int]]] = defaultdict(
        lambda: {"accepted": defaultdict(int), "confirmed": defaultdict(int)}
    )

    campus_map = {"SAV": "Savannah", "ATL": "Atlanta", "ELN": "SCADnow"}

    for sheet_name in wb.sheetnames:
        match = SHEET_RE.match(sheet_name.strip())
        if not match:
            continue
        term = match.group("term")
        campus_code = match.group("campus")
        stype = match.group("stype")
        if term not in terms:
            continue
        campus_name = campus_map.get(campus_code, campus_code)

        ws = wb[sheet_name]

        header_row_idx = None
        headers = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), start=1):
            if row and ("1st Interest" in row or "Major" in row):
                header_row_idx = i
                headers = list(row)
                break
        if not header_row_idx or not headers:
            continue

        def col_index(label: str) -> Optional[int]:
            try:
                return headers.index(label)
            except ValueError:
                return None

        interest_idx = col_index("1st Interest")
        major_idx = col_index("Major")
        decision_idx = col_index("Latest Decision")
        mat_fee_date_idx = col_index("MAT Fee Paid Date")

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row or not any(row):
                continue

            raw_interest = row[interest_idx] if (interest_idx is not None and interest_idx < len(row)) else None
            raw_major = row[major_idx] if (major_idx is not None and major_idx < len(row)) else None
            raw_decision = row[decision_idx] if (decision_idx is not None and decision_idx < len(row)) else None
            raw_mat_fee_date = row[mat_fee_date_idx] if (mat_fee_date_idx is not None and mat_fee_date_idx < len(row)) else None

            interest = raw_interest or raw_major
            if not interest:
                continue
            interest_norm = normalize_program_base(str(interest))
            if not interest_norm:
                continue

            counts[(term, campus_name, stype)]["accepted"][interest_norm] += 1

            decision_text = normalize_text(str(raw_decision)) if raw_decision else ""
            is_confirmed = bool(raw_mat_fee_date) or ("MAT FEE" in decision_text)
            if is_confirmed:
                counts[(term, campus_name, stype)]["confirmed"][interest_norm] += 1

    return counts


def add_demand(
    demand: dict[tuple[str, str], float],
    course: str,
    campus: str,
    seats: float,
) -> None:
    if seats <= 0:
        return
    demand[(course, campus)] += seats


def forecast_spring26(
    program_index: dict[str, list[SequenceRow]],
    admissions_counts: dict[tuple[str, str, str], dict[str, dict[str, int]]],
    scenario: str,
    progression_rate: float,
    include_transfers: bool,
    include_second_year: bool,
    fall24_total_cohort: float,
) -> tuple[dict[tuple[str, str], float], list[tuple[tuple[str, str, str], str, float]]]:
    """
    Returns:
      demand[(course, campus)] = projected seats
      unmapped = [((term, campus, cohort_year), interest, count)]
    """
    demand: dict[tuple[str, str], float] = defaultdict(float)
    unmapped: list[tuple[tuple[str, str, str], str, float]] = []

    # Cohort-to-Spring26 mapping: which quarter plan to apply.
    # - Fall 2025 starters (202610): Spring 2026 is "spring" quarter of the map (2 transitions).
    # - Winter 2026 starters (202620): Spring 2026 is "spring" quarter of the map (1 transition).
    # - Spring 2026 starters (202630): Spring 2026 is first term; use "fall" quarter of the map (0 transitions).
    cohort_rules = {
        # term: (year_label, quarter_key, transitions_to_spring26)
        "202610": ("first year", "spring", 2),
        "202620": ("first year", "spring", 1),
        "202630": ("first year", "fall", 0),
    }

    def apply_counts(
        *,
        cohort_key: tuple[str, str, str],  # (term, campus, cohort_year_label)
        year_label: str,
        quarter_key: str,
        multiplier: float,
        counts_by_interest: dict[str, float],
    ) -> None:
        term, campus, cohort_year = cohort_key
        for interest, count in counts_by_interest.items():
            if count <= 0:
                continue

            base_keys = best_program_matches(program_index, interest)
            if not base_keys:
                unmapped.append(((term, campus, cohort_year), interest, count))
                continue

            # Expand to matching year rows
            matching_rows: list[SequenceRow] = []
            for base_key in base_keys:
                for row in program_index.get(base_key, []):
                    if row.year.strip().lower() != year_label:
                        continue
                    if campus_matches(row, campus):
                        matching_rows.append(row)

            if not matching_rows:
                # Fallback: ignore campus constraint (still same year)
                for base_key in base_keys:
                    for row in program_index.get(base_key, []):
                        if row.year.strip().lower() != year_label:
                            continue
                        matching_rows.append(row)

            if not matching_rows:
                unmapped.append(((term, campus, cohort_year), interest, count))
                continue

            per_row = (count * multiplier) / len(matching_rows)
            for row in matching_rows:
                quarter_plan: QuarterPlan = getattr(row, quarter_key)
                if not quarter_plan.courses:
                    continue
                if quarter_plan.mode == "choice":
                    per_course = per_row / len(quarter_plan.courses) if quarter_plan.courses else 0
                    for course in quarter_plan.courses:
                        add_demand(demand, course, campus, per_course)
                else:
                    for course in quarter_plan.courses:
                        add_demand(demand, course, campus, per_row)

    # 1) First-year Spring 2026 demand driven by admissions terms.
    for (term, campus, stype), scenarios in admissions_counts.items():
        if term not in cohort_rules:
            continue
        if (not include_transfers) and stype == "TR":
            continue

        year_label, quarter_key, transitions = cohort_rules[term]
        multiplier = progression_rate**transitions
        apply_counts(
            cohort_key=(term, campus, "first year"),
            year_label=year_label,
            quarter_key=quarter_key,
            multiplier=multiplier,
            counts_by_interest={k: float(v) for k, v in scenarios[scenario].items()},
        )

    # 2) Second-year Spring 2026 demand (approx): estimate Fall 2024 cohort distribution using Fall 2025
    # freshman distribution, scaled to a provided Fall 2024 cohort size.
    if include_second_year and fall24_total_cohort > 0:
        # Build Fall 2025 freshman distribution by campus from admissions term 202610.
        fall25_fr_by_campus: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
        fall25_fr_totals: dict[str, float] = defaultdict(float)

        for (term, campus, stype), scenarios in admissions_counts.items():
            if term != "202610" or stype != "FR":
                continue
            for interest, count in scenarios[scenario].items():
                fall25_fr_by_campus[campus][interest] += float(count)
                fall25_fr_totals[campus] += float(count)

        total_fall25_fr = sum(fall25_fr_totals.values())
        if total_fall25_fr > 0:
            # Allocate Fall 2024 cohort total to campuses proportional to Fall 2025 freshman distribution.
            fall24_by_campus_total = {
                campus: (fall25_fr_totals[campus] / total_fall25_fr) * float(fall24_total_cohort)
                for campus in fall25_fr_totals.keys()
            }

            # From Fall 2024 start to Spring 2026 is 6 term-to-term transitions.
            second_year_multiplier = progression_rate**6

            for campus, campus_total in fall24_by_campus_total.items():
                if campus_total <= 0 or fall25_fr_totals.get(campus, 0) <= 0:
                    continue
                campus_dist = fall25_fr_by_campus[campus]
                for interest, fall25_count in campus_dist.items():
                    share = fall25_count / fall25_fr_totals[campus] if fall25_fr_totals[campus] else 0
                    est_fall24_interest_count = campus_total * share
                    apply_counts(
                        cohort_key=("202510", campus, "second year"),
                        year_label="second year",
                        quarter_key="spring",
                        multiplier=second_year_multiplier,
                        counts_by_interest={interest: est_fall24_interest_count},
                    )

    return dict(demand), unmapped


def write_forecast_csv(
    output_path: Path,
    demand: dict[tuple[str, str], float],
    capacity: int,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = []
    for (course, campus), seats in demand.items():
        sections = int(math.ceil(seats / capacity)) if seats > 0 else 0
        rows.append((course, campus, seats, sections))
    rows.sort(key=lambda r: (r[0], r[1]))

    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["course", "campus", "projected_seats", "sections"])
        for course, campus, seats, sections in rows:
            writer.writerow([course, campus, f"{seats:.2f}", sections])


def print_summary(demand: dict[tuple[str, str], float], capacity: int) -> None:
    by_course = defaultdict(float)
    for (course, _campus), seats in demand.items():
        by_course[course] += seats

    print("\nSpring 2026 forecast (all campuses combined)")
    print(f"{'Course':<10}  {'Seats':>10}  {'Sections':>10}")
    print("-" * 34)
    for course in sorted(by_course.keys()):
        seats = by_course[course]
        sections = int(math.ceil(seats / capacity)) if seats > 0 else 0
        print(f"{course:<10}  {seats:>10.2f}  {sections:>10}")

    print("\nSpring 2026 forecast by campus")
    print(f"{'Course':<10}  {'Campus':<8}  {'Seats':>10}  {'Sections':>10}")
    print("-" * 44)
    for (course, campus) in sorted(demand.keys()):
        seats = demand[(course, campus)]
        sections = int(math.ceil(seats / capacity)) if seats > 0 else 0
        print(f"{course:<10}  {campus:<8}  {seats:>10.2f}  {sections:>10}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Forecast Spring 2026 FOUN sections using the sequencing map by major.")
    parser.add_argument("--sequence-map", default="Data/FOUN_sequencing_map_by_major.csv")
    parser.add_argument("--admissions", default="Data/PZSAAPF-SL31 - Accepted Applicants with Latest Decision.xlsx")
    parser.add_argument("--historical", default="Data/FOUN_Historical.csv", help="Used to approximate Fall 2024 cohort size for second-year demand.")
    parser.add_argument("--scenario", choices=["accepted", "confirmed"], default="confirmed")
    parser.add_argument("--progression-rate", type=float, default=0.95, help="Per-term retention/progression rate applied to prior cohorts.")
    parser.add_argument("--capacity", type=int, default=20, help="Students per section.")
    parser.add_argument("--include-transfers", action="store_true", help="Include transfer admits in the demand model.")
    parser.add_argument("--exclude-second-year", action="store_true", help="Exclude estimated second-year Spring 2026 demand.")
    parser.add_argument("--fall24-total-cohort", type=float, default=0.0, help="Override estimated Fall 2024 cohort size used for second-year demand.")
    parser.add_argument("--output", default="Data/FOUN_Spring26_Section_Forecast.csv")
    args = parser.parse_args()

    sequence_rows = load_sequence_rows(Path(args.sequence_map))
    program_index = build_program_index(sequence_rows)

    admissions_counts = parse_admissions_counts(Path(args.admissions), terms={"202610", "202620", "202630"})

    fall24_total = float(args.fall24_total_cohort)
    if fall24_total <= 0 and not args.exclude_second_year:
        # Approximate Fall 2024 cohort size from historical term 202510 using the larger of DRAW 100 vs DSGN 100.
        fall24_draw100 = 0.0
        fall24_dsgn100 = 0.0
        with Path(args.historical).open(newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                try:
                    term = int(row.get("TERM") or 0)
                except ValueError:
                    continue
                if term != 202510:
                    continue
                subj = (row.get("SUBJ") or "").strip().upper()
                crs = (row.get("CRS NUMBER") or "").strip()
                try:
                    act_enr = float(row.get("ACT ENR") or 0)
                except ValueError:
                    act_enr = 0.0
                if subj == "DRAW" and crs == "100":
                    fall24_draw100 += act_enr
                elif subj == "DSGN" and crs == "100":
                    fall24_dsgn100 += act_enr
        fall24_total = max(fall24_draw100, fall24_dsgn100)

    demand, unmapped = forecast_spring26(
        program_index=program_index,
        admissions_counts=admissions_counts,
        scenario=args.scenario,
        progression_rate=args.progression_rate,
        include_transfers=args.include_transfers,
        include_second_year=not args.exclude_second_year,
        fall24_total_cohort=fall24_total,
    )

    write_forecast_csv(Path(args.output), demand=demand, capacity=args.capacity)
    print_summary(demand=demand, capacity=args.capacity)

    if unmapped:
        print("\nUnmapped interests (top 20 by count)")
        unmapped_sorted = sorted(unmapped, key=lambda x: x[2], reverse=True)
        for (src_term, src_campus, cohort_year_label), interest, count in unmapped_sorted[:20]:
            print(f"{src_term} {src_campus:<8} {cohort_year_label:<11} {count:>6.1f}  {interest}")

    print(f"\nWrote {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
