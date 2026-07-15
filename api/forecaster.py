"""
Sequence-based FOUN enrollment forecasting.

Pure functions extracted from forecast_spring26_from_sequence_guides.py
so the FastAPI backend can call them without argparse/sys.exit coupling.

Generalized to forecast any target quarter (Spring, Summer, Fall, Winter).
"""

import csv
import logging
import math
import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple, DefaultDict

FOUN_CODE_RE = re.compile(r"\bFOUN\s*(\d{3})\b", re.IGNORECASE)

# Quarter cycle: each quarter's two feeders in order (closer, farther)
QUARTER_CYCLE = {
    "spring": ("winter", "fall"),
    "summer": ("spring", "winter"),
    "fall":   ("summer", "spring"),
    "winter": ("fall", "summer"),
}

# SCAD term code quarter digits
QUARTER_CODES = {"fall": 10, "winter": 20, "spring": 30, "summer": 40}

# FOUN curriculum launched Fall 2025 (AY 2025-2026).  Each subsequent AY
# adds one more cohort year until all four years are represented.
_FOUN_CURRICULUM_START = 2025
_YEAR_LABELS = ["First Year", "Second Year", "Third Year", "Fourth Year"]

# Crosswalk: Cognos major code → list of exact sequencing-map program names.
# When a code maps to multiple programs (e.g. ANIM → 3 animation tracks),
# enrollment is split evenly.  Codes absent from this dict are silently ignored.
_COGNOS_TO_SEQ_PROGRAMS: Dict[str, List[str]] = {
    "ACCE": ["ACCESSORY DESIGN"],
    "ACT":  ["ACTING Group A"],
    "ADBR": ["ADVERTISING AND BRANDING"],
    "ANIM": [
        "Fall Only Winter Only Spring Only ANIMATION 2D Animation",
        "Fall Only Winter Only Spring Only ANIMATION 3D Character Animation",
        "Fall Only Winter Only Spring Only ANIMATION Storytelling and Concept Development",
    ],
    "ARCH": ["ARCHITECTURE"],
    "ARLH": ["ARCHITECTURAL HISTORY"],
    "ARTH": ["ART HISTORY"],
    "ARVR": ["Fall Only Winter Only Spring Only IMMERSIVE REALITY"],
    "DWRI": ["DRAMATIC WRITING"],
    "EQST": ["EQUESTRIAN STUDIES"],
    "FASH": ["FASHION Group A"],
    "FASM": ["FASHION MARKETING AND MANAGEMENT"],
    "FIBR": ["Fall Only Winter Only Spring On ly FIBERS"],  # typo preserved from seq map
    "FILM": [
        "FILM AND TELEVISION Group A",
        "FILM AND TELEVISION Group B",
        "FILM AND TELEVISION Group C",
    ],
    "FURN": ["Fall Only Winter Only Spring Only FURNITURE DESIGN"],
    "GAME": ["GAME DEVELOPMENT", "Fall Only Winter Only Spring Only GAME DEVELOPMENT"],
    "GRDS": ["GRAPHIC DESIGN"],
    "IDUS": ["INTERIOR DESIGN"],
    "ILLU": [
        "ILLUSTRATION",
        "ILLUSTRATION Dynamic Illustration and Publication Design, Group A",
        "ILLUSTRATION Illustration for Surface Design, Group B",
        "ILLUSTRATION Visual Development, Group B",
    ],
    "INDS": ["INDUSTRIAL DESIGN"],
    "JEWL": ["Fall Only Winter Only Spring Only JEWELRY"],
    "MOME": ["MOTION MEDIA DESIGN"],
    "PHOT": ["PHOTOGRAPHY"],
    "PNTG": ["PAINTING"],
    "PROD": [
        "PRODUCTION DESIGN Costume Design, Group A",
        "PRODUCTION DESIGN Lighting Design, Group A",
        "PRODUCTION DESIGN Set Design and Art Direction Design, Group A",
    ],
    "SEQA": ["SEQUENTIAL ART", "SEQUENTIAL ART Group A"],
    "SERV": ["SERVICE DESIGN"],
    "SNDS": ["SOUND DESIGN"],
    "SOCL": ["SOCIAL STRATEGY AND MANAGEMENT"],
    "UXDG": ["USER EXPERIENCE DESIGN (UXD)"],
    "UXR":  ["USER EXPERIENCE RESEARCH (UXR)"],
    "VFX":  ["VISUAL EFFECTS AND TECHNICAL ANIMATION"],
    "VSFX": ["VISUAL EFFECTS AND TECHNICAL ANIMATION"],
    "WRIT": ["Fall Only Winter Only Spring Only WRITING"],
}


def _active_curriculum_years(term_code: str) -> Optional[List[str]]:
    """Return curriculum year labels active for the given SCAD term code.

    SCAD codes are YYYYQQ.  The FOUN curriculum was rolled out simultaneously
    across all four cohort years in Fall 2025 (term code 202610), not phased
    in cohort-by-cohort.  This is confirmed by PZSMSCP Spring 2026 actuals:
    Second/Third Year FOUN courses (FOUN 250/251/260/330/331/360) carry real
    ACT enrollment that cannot exist if only First-Year cohorts had entered
    the curriculum.

    For any term at or after Fall 2025, return all four cohort labels.
    For pre-curriculum terms (before Fall 2025), return ``None`` so the
    year filter is disabled and historical backtests pass all seq-map rows
    through unfiltered.

    Returning the explicit four-label list (rather than ``None``) for
    post-launch terms has identical effect on ``load_sequence_mappings``
    because the seq map only contains First/Second/Third/Fourth Year rows;
    the explicit form documents intent and protects against future seq-map
    additions of non-cohort year labels.

    See ``docs/SPRING_2026_BACKTEST.md`` for the evidence behind this change
    and the prior (incorrect) phased-launch assumption.
    """
    # Fall 2025 = AY 2025-2026 = term code 202610.  Term codes are YYYYQQ
    # where YYYY is the academic year start +1 for fall, so 202610 is the
    # first post-launch term.  Anything before that (e.g. 202510 Fall 2024,
    # 202530 Spring 2025) returns None to disable the filter for backtests.
    yyyy = int(str(term_code)[:4])
    if yyyy <= _FOUN_CURRICULUM_START:
        return None
    return list(_YEAR_LABELS)


def resolve_term_info(target_term: str) -> Dict:
    """Parse a human-readable term like 'Summer 2026' into quarter info
    with feeder term codes.

    Returns dict with:
        target_quarter, target_term_code,
        closer_feeder: {quarter, term_code, multiplier_exp},
        farther_feeder: {quarter, term_code, multiplier_exp}
    """
    target_term = target_term.strip()
    # Accept YYYYQQ numeric term codes (e.g. "202630") and convert to "Quarter YYYY"
    if re.match(r'^\d{6}$', target_term):
        code_year = int(target_term[:4])
        code_suffix = int(target_term[4:])
        suffix_to_quarter = {10: "fall", 20: "winter", 30: "spring", 40: "summer"}
        if code_suffix not in suffix_to_quarter:
            raise ValueError(f"Invalid term code suffix '{code_suffix}' in '{target_term}'. Expected 10/20/30/40.")
        q = suffix_to_quarter[code_suffix]
        cal_year = code_year - 1 if code_suffix == 10 else code_year
        target_term = f"{q.capitalize()} {cal_year}"

    parts = target_term.split()
    if len(parts) != 2:
        raise ValueError(f"Invalid target_term format: '{target_term}'. Expected 'Quarter YYYY'.")
    quarter_name = parts[0].lower()
    calendar_year = int(parts[1])

    if quarter_name not in QUARTER_CYCLE:
        raise ValueError(f"Unknown quarter: '{quarter_name}'. Must be spring/summer/fall/winter.")

    # Academic year: Fall uses next calendar year's academic code
    if quarter_name == "fall":
        academic_year = calendar_year + 1
    else:
        academic_year = calendar_year

    target_term_code = str(academic_year) + str(QUARTER_CODES[quarter_name])

    closer_q, farther_q = QUARTER_CYCLE[quarter_name]

    def _feeder_term_code(feeder_quarter: str) -> str:
        """Compute the term code for a feeder quarter preceding the target."""
        # Walk backwards: the closer feeder is 1 quarter before target,
        # the farther feeder is 2 quarters before target.
        # We need to figure out the calendar year for each feeder.
        # Determine the calendar year of the feeder
        # The feeder is in the same academic year cycle or the previous one
        if feeder_quarter == "fall":
            # Fall always uses calendar_year of the fall itself
            # Fall before winter/spring/summer of calendar_year is Fall (calendar_year - 1)
            feeder_cal_year = calendar_year - 1
            feeder_acad_year = feeder_cal_year + 1  # Fall academic year
        elif feeder_quarter == "winter":
            if quarter_name == "fall":
                # Winter before Fall: same calendar year
                feeder_cal_year = calendar_year
            else:
                # Winter before Spring/Summer of same year
                feeder_cal_year = calendar_year
            feeder_acad_year = feeder_cal_year
        elif feeder_quarter == "spring":
            feeder_cal_year = calendar_year
            feeder_acad_year = feeder_cal_year
        else:  # summer
            if quarter_name == "fall":
                feeder_cal_year = calendar_year
            elif quarter_name == "winter":
                # Winter 2027 ← Summer 2026
                feeder_cal_year = calendar_year - 1
            else:
                feeder_cal_year = calendar_year
            feeder_acad_year = feeder_cal_year

        # Apply SCAD academic year convention: Fall uses next year
        if feeder_quarter == "fall":
            feeder_acad_year = feeder_cal_year + 1

        return str(feeder_acad_year) + str(QUARTER_CODES[feeder_quarter])

    return {
        "target_quarter": quarter_name,
        "target_term_code": target_term_code,
        "closer_feeder": {
            "quarter": closer_q,
            "term_code": _feeder_term_code(closer_q),
            "multiplier_exp": 1,
        },
        "farther_feeder": {
            "quarter": farther_q,
            "term_code": _feeder_term_code(farther_q),
            "multiplier_exp": 2,
        },
    }


def normalize_text(value: str) -> str:
    value = value or ""
    value = str(value).upper()
    value = value.replace("&", " AND ")
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def parse_campuses(campus_raw: str) -> Tuple[str, ...]:
    campus_norm = normalize_text(campus_raw)
    if not campus_norm:
        return tuple()
    campus_norm = campus_norm.replace("MAJOR COURSE SEQUENCING GUIDE", "").strip()
    if campus_norm == "GENERAL":
        return ("GENERAL",)
    parts = [p.strip() for p in campus_norm.split("|")]
    parts = [p for p in parts if p]
    return tuple(parts)


def campus_matches(campuses: Iterable[str], campus: str) -> bool:
    if "GENERAL" in campuses:
        return True
    return campus in campuses


def extract_foun_codes(cell_value: str) -> List[str]:
    if not cell_value:
        return []
    seen = set()
    codes = []
    for match in FOUN_CODE_RE.findall(str(cell_value)):
        code = f"FOUN {match}"
        if code not in seen:
            seen.add(code)
            codes.append(code)
    return codes


def parse_quarter_courses(cell_value: str) -> List[Tuple[str, float]]:
    text = str(cell_value or "").strip()
    if not text:
        return []
    courses = extract_foun_codes(text)
    if not courses:
        return []
    is_choice = "CHOICE" in text.upper()
    weight = 1.0 / len(courses) if is_choice else 1.0
    return [(course, weight) for course in courses]


def _select_anchor_courses(
    courses: List[Tuple[str, float]],
    freq: Dict[str, int],
) -> List[Tuple[str, float]]:
    """Pick one representative from concurrent courses to avoid double-counting.

    Concurrent courses (weight >= 1.0) are co-requisites taken by the same
    cohort. Picking only the most-frequent one prevents counting that cohort
    multiple times. CHOICE courses (weight < 1.0) are alternatives — kept as-is.
    """
    if not courses:
        return courses
    concurrent = [(c, w) for c, w in courses if w >= 1.0]
    if len(concurrent) <= 1:
        return courses
    best = max(concurrent, key=lambda cw: freq.get(cw[0], 0))
    return [best]


def load_sequence_mappings(
    path: Path,
    target_quarter: str,
    closer_quarter: str,
    farther_quarter: str,
    year_filter: Optional[List[str]] = None,
    enrollment_weights: Optional[Dict[str, Dict[str, Dict[str, float]]]] = None,
) -> Dict[str, Dict[str, Dict[str, float]]]:
    """Load the sequencing map CSV and build mappings for the given quarters.

    Args:
        year_filter: When provided, only rows whose ``year`` column value is in
            this list are included.  Pass ``_active_curriculum_years(term_code)``
            to automatically exclude cohort years that don't exist yet.

    Returns per-campus dicts with keys:
        farther_to_target     – source→target weights, only from rows with no closer course
        farther_source_totals – total program weight per farther source across ALL rows
                                (used to normalize farther-feeder proportions correctly)
        closer_to_target      – source→target weights from rows with a closer course
        closer_source_totals  – total program weight per closer source, computed from
                                UNFILTERED data so co-occurring courses get diluted
        target_counts
    """
    mappings = {
        "SAVANNAH": {
            "farther_to_target": defaultdict(float),
            "farther_source_totals": defaultdict(float),
            "closer_to_target": defaultdict(float),
            "closer_source_totals": defaultdict(float),
            "target_counts": defaultdict(float),
        },
        "SCADNOW": {
            "farther_to_target": defaultdict(float),
            "farther_source_totals": defaultdict(float),
            "closer_to_target": defaultdict(float),
            "closer_source_totals": defaultdict(float),
            "target_counts": defaultdict(float),
        },
        "ATLANTA": {
            "farther_to_target": defaultdict(float),
            "farther_source_totals": defaultdict(float),
            "closer_to_target": defaultdict(float),
            "closer_source_totals": defaultdict(float),
            "target_counts": defaultdict(float),
        },
    }

    # --- Pass 1: Count how often each course appears as a concurrent source ---
    # This determines which course is the best anchor per quarter.
    closer_freq: Dict[str, int] = defaultdict(int)
    farther_freq: Dict[str, int] = defaultdict(int)

    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if year_filter is not None and row.get("year", "").strip() not in year_filter:
                continue
            for c, w in parse_quarter_courses(row.get(closer_quarter)):
                if w >= 1.0:
                    closer_freq[c] += 1
            for c, w in parse_quarter_courses(row.get(farther_quarter)):
                if w >= 1.0:
                    farther_freq[c] += 1

    # --- Pass 2: Build mappings using anchor-filtered source courses ---
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if year_filter is not None and row.get("year", "").strip() not in year_filter:
                continue
            campuses = parse_campuses(row.get("campus", ""))
            if not campuses:
                continue

            farther_raw = parse_quarter_courses(row.get(farther_quarter))
            closer_raw = parse_quarter_courses(row.get(closer_quarter))
            target_courses = parse_quarter_courses(row.get(target_quarter))

            farther_courses = _select_anchor_courses(farther_raw, farther_freq)
            closer_courses = _select_anchor_courses(closer_raw, closer_freq)

            for campus in ("SAVANNAH", "SCADNOW", "ATLANTA"):
                if not campus_matches(campuses, campus):
                    continue

                # Compute enrollment-based row weight (1.0 if no weights provided)
                row_weight = 1.0
                if enrollment_weights:
                    program = row.get("program", "").strip().upper()
                    campus_w = enrollment_weights.get(campus, {})
                    # Use closer feeder course enrollment if available, else farther
                    if closer_raw:
                        row_weight = campus_w.get(closer_raw[0][0], {}).get(program, 0.0)
                    elif farther_raw:
                        row_weight = campus_w.get(farther_raw[0][0], {}).get(program, 0.0)

                # Source totals accumulate for ALL rows, including those with no Spring
                # target.  Programs that take a feeder course in Winter but don't route
                # to any Spring FOUN course (e.g. Animation students whose CHOICE winter
                # FOUN has an empty Spring, or Motion Media Design taking FOUN 251 in
                # Winter with nothing in Spring) must still dilute the fraction so that
                # the few programs which do have Spring targets don't claim 100% of that
                # feeder's enrollment.  Without this, source_totals[FOUN 251] = 1.0
                # (Photography only) and all 264 Winter FOUN 251 students project to
                # FOUN 220, even though Fashion Marketing and Motion Media Design also
                # fill those seats and go nowhere in Spring.
                # Each program-row contributes its weight ONCE per source course,
                # regardless of how many target courses it routes to.  A CHOICE
                # target cell (weights summing to 1.0) and a single target are one
                # cohort; so is a co-requisite cell ("FOUN 240; FOUN 245", weights
                # summing to k) whose students take ALL listed targets.  Multiplying
                # by the target-weight sum here would split co-req demand 1/k and
                # over-dilute sibling programs sharing the same feeder.
                for closer_course, closer_weight in closer_raw:
                    mappings[campus]["closer_source_totals"][closer_course] += closer_weight * row_weight
                # Accumulate from the UNFILTERED farther courses, mirroring the
                # closer side: a non-anchor co-requisite still dilutes its own
                # denominator, because its cohort is routed via the anchor.
                for farther_course, farther_weight in farther_raw:
                    mappings[campus]["farther_source_totals"][farther_course] += farther_weight * row_weight
                if not target_courses:
                    continue  # nothing else to build for this campus row

                for target_course, target_weight in target_courses:
                    mappings[campus]["target_counts"][target_course] += target_weight * row_weight

                # Only populate farther_to_target for rows that have no closer-quarter course.
                # When a row has both a closer course and a farther course leading to the
                # same target, those students will be captured via the closer feeder path.
                # Adding the farther feeder as well would double-count the same cohort.
                if not closer_courses:
                    for farther_course, farther_weight in farther_courses:
                        for target_course, target_weight in target_courses:
                            key = (farther_course, target_course)
                            mappings[campus]["farther_to_target"][key] += farther_weight * target_weight * row_weight

                for closer_course, closer_weight in closer_courses:
                    for target_course, target_weight in target_courses:
                        key = (closer_course, target_course)
                        mappings[campus]["closer_to_target"][key] += closer_weight * target_weight * row_weight

    return mappings


def parse_number(value: str) -> float:
    if value is None:
        return 0.0
    text = str(value).replace(",", "").strip()
    if text == "":
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def normalize_demand_metric(metric: Optional[str]) -> str:
    """Return a supported enrollment metric name.

    ``actual`` is the default because it matches the existing behavior.
    ``max`` lets planners forecast scheduled seats instead of enrolled seats.
    ``actual_plus_waitlist`` treats waitlisted students as constrained demand.
    """
    value = (metric or "actual").strip().lower()
    if value in {"actual", "act", "act_enr"}:
        return "actual"
    if value in {"max", "max_enr", "scheduled", "capacity"}:
        return "max"
    if value in {"actual_plus_waitlist", "act_plus_waitlist", "waitlist", "demand"}:
        return "actual_plus_waitlist"
    return "actual"


def _metric_enrollment(getter, metric: str) -> float:
    """Read the selected demand metric from a row-like getter.

    ``metric`` must already be normalized via ``normalize_demand_metric``;
    loaders normalize once per file, not once per row.
    """
    if metric == "max":
        value = getter("MAX ENR")
        # CSV blank cells arrive as "" (xlsx blanks arrive as None); both mean
        # "no cap recorded" and must fall back to actual enrollment, not 0.
        if value is None or str(value).strip() == "":
            value = getter("ACT ENR")
        return parse_number(value)
    if metric == "actual_plus_waitlist":
        return parse_number(getter("ACT ENR")) + parse_number(getter("WL ACT ENR"))
    return parse_number(getter("ACT ENR"))


def load_crosswalk(crosswalk_path: Path) -> Dict[str, str]:
    """Load legacy→FOUN course code mappings from the crosswalk CSV.

    Returns a dict like {"DSGN 100": "FOUN 110", "DRAW 200": "FOUN 230", ...}.
    Returns an empty dict if the file doesn't exist.
    """
    if not crosswalk_path.is_file():
        return {}
    mapping: Dict[str, str] = {}
    with crosswalk_path.open(newline="", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.DictReader(f)
        for row in reader:
            legacy = (row.get("legacy_code") or "").strip()
            foun = (row.get("foun_code") or "").strip()
            if legacy and foun:
                mapping[legacy] = foun
    return mapping


def _load_admits_rows(path: Path) -> Tuple[List[Tuple], Dict[str, int]]:
    """Return data rows and useful column indices from a PZSAAPF xlsx file."""
    if not path.is_file():
        return [], {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception:
        return [], {}

    header: Dict[str, int] = {}
    data_start = 0
    for idx, row in enumerate(rows):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if "student id" in cells:
            for i, cell in enumerate(cells):
                if cell:
                    header[cell] = i
            data_start = idx + 1
            break

    if not header:
        # Legacy fallback: column M is campus and column U is registered courses.
        header = {"campus": 12, "currently registered courses (no wl)": 20}
        data_start = 0

    return rows[data_start:], header


def _admit_header_index(header: Dict[str, int], *names: str, fallback: int) -> int:
    for name in names:
        if name in header:
            return header[name]
    return fallback


_ADMIT_CAMPUS_MAP = {"M": "SAVANNAH", "O": "SCADNOW"}


def _decode_admit_rows(path: Path) -> Optional[List[Tuple[str, str]]]:
    """Return [(campus, registered_courses_text), ...] for decodable admit rows.

    One row decoder shared by ``load_admits_foun_demand`` and
    ``analyze_admits_registration``, so a Cognos column rename is fixed once
    and both always read the same columns. Returns None when the file is
    missing or unreadable; rows with campus codes outside the map are skipped.
    """
    rows, header = _load_admits_rows(path)
    if not rows:
        return None
    campus_idx = _admit_header_index(header, "campus", fallback=12)
    courses_idx = _admit_header_index(
        header,
        "currently registered courses (no wl)",
        "currently registered courses",
        fallback=20,
    )
    decoded: List[Tuple[str, str]] = []
    for row in rows:
        if not row or row[0] is None:
            continue
        campus_code = str(row[campus_idx]).strip() if len(row) > campus_idx and row[campus_idx] else ""
        campus = _ADMIT_CAMPUS_MAP.get(campus_code)
        if not campus:
            continue
        reg_courses = row[courses_idx] if len(row) > courses_idx else None
        decoded.append((campus, str(reg_courses or "").strip()))
    return decoded


def load_admits_foun_demand(path: Path) -> Dict[str, Dict[str, int]]:
    """Extract FOUN course demand from accepted applicants xlsx (PZSAAPF-SL31).

    Reads the "Currently Registered Courses (NO WL)" column for each student.
    Campus M → SAVANNAH, O → SCADNOW. Returns {campus: {foun_course: count}}.
    Returns empty dict if file is missing or unreadable.
    """
    decoded = _decode_admit_rows(path)
    if decoded is None:
        return {}

    result: Dict[str, DefaultDict] = {
        "SAVANNAH": defaultdict(int),
        "SCADNOW": defaultdict(int),
    }
    for campus, reg_text in decoded:
        if not reg_text:
            continue
        for code in FOUN_CODE_RE.findall(reg_text):
            result[campus][f"FOUN {code}"] += 1

    return {k: dict(v) for k, v in result.items()}


def analyze_admits_registration(path: Path) -> Dict:
    """Return registration maturity diagnostics for an admits report.

    The admits report is only useful for intro-course demand once enough admits
    have registered. A low FOUN-registration rate usually means the snapshot was
    pulled too early, so FOUN 110/111 will be undercounted.
    """
    decoded = _decode_admit_rows(path)
    if decoded is None:
        return {}

    by_campus: Dict[str, Dict[str, float]] = {
        "SAVANNAH": {"total": 0, "registered_any": 0, "registered_foun": 0},
        "SCADNOW": {"total": 0, "registered_any": 0, "registered_foun": 0},
    }

    for campus, reg_text in decoded:
        by_campus[campus]["total"] += 1
        if reg_text:
            by_campus[campus]["registered_any"] += 1
        if FOUN_CODE_RE.search(reg_text):
            by_campus[campus]["registered_foun"] += 1

    warnings: List[str] = []
    for campus, stats in by_campus.items():
        total = stats["total"]
        if total <= 0:
            continue
        any_rate = stats["registered_any"] / total
        foun_rate = stats["registered_foun"] / total
        stats["registered_any_rate"] = round(any_rate, 3)
        stats["registered_foun_rate"] = round(foun_rate, 3)
        if total >= 20 and foun_rate < 0.50:
            label = CAMPUS_DISPLAY.get(campus, campus)
            warnings.append(
                f"Admits snapshot may be early for {label}: only {foun_rate:.0%} "
                f"of admits are registered for a FOUN course. FOUN 110/111 demand may be low."
            )

    return {"by_campus": by_campus, "warnings": warnings}


def load_enrollment_by_major(path: Path) -> Dict[str, Dict[str, Dict[str, float]]]:
    """Load Cognos enrollment-by-major report (xlsx or CSV).

    **DISABLED IN PRODUCTION** as of May 2026: the end user lacks access to
    student-level Cognos reports needed to produce the by-major aggregated
    input this loader expects. ``forecast_config.json`` ships with
    ``"enrollmentByMajorFile": null``, which short-circuits invocation in
    ``run_sequence_forecast``. The function is preserved for development and
    testing use, and may be re-enabled if/when an aggregated data source
    (e.g. an institutional research extract that does not require
    student-level access) becomes available.

    **xlsx (Cognos export):**
      - Finds the header row by scanning for a cell containing 'term'
      - Course format ``FOUN110`` is normalised to ``FOUN 110``
      - Major column contains Cognos abbreviated codes (ARCH, ANIM, …);
        these are translated to seq-map program names via ``_COGNOS_TO_SEQ_PROGRAMS``
      - Codes absent from the crosswalk are silently ignored
      - One-to-many codes (e.g. ANIM → 3 tracks) split enrollment evenly

    **CSV (legacy / test format):**
      - Expected columns: term, course, campus, major, enrollment
      - Major values must already be seq-map program names (uppercased on read)

    Campus: SAV → SAVANNAH, NOW → SCADNOW, ATL → ATLANTA
    Returns {campus: {foun_course: {program_upper: total_enrollment}}}.
    Returns empty dict if file is missing.
    """
    if not path.is_file():
        return {}

    CAMPUS_MAP = {"SAV": "SAVANNAH", "NOW": "SCADNOW", "ATL": "ATLANTA"}
    result: Dict[str, Dict[str, Dict[str, float]]] = {}
    # De-duplicate dropped-code warnings within a single load: log each
    # unmapped Cognos code at most once per call.
    _warned_unmapped: set = set()

    def _accumulate(campus_raw, course_raw, major_raw, enroll_raw, use_crosswalk: bool) -> None:
        campus = CAMPUS_MAP.get(str(campus_raw).strip().upper())
        if not campus:
            return
        m = re.match(r"^FOUN\s*(\d{3})$", str(course_raw).strip(), re.IGNORECASE)
        if not m:
            return
        course = f"FOUN {m.group(1)}"
        try:
            enroll = float(enroll_raw) if enroll_raw is not None else 0.0
        except (ValueError, TypeError):
            return
        if use_crosswalk:
            code = str(major_raw).strip().upper()
            programs = _COGNOS_TO_SEQ_PROGRAMS.get(code)
            if not programs:
                if code and code not in _warned_unmapped:
                    logging.warning(
                        "Unmapped Cognos major code %r dropped from enrollment weights "
                        "(first seen on %s, enrollment=%s); add to _COGNOS_TO_SEQ_PROGRAMS to include.",
                        code, course, enroll,
                    )
                    _warned_unmapped.add(code)
                return
            per_prog = enroll / len(programs)
            for prog in programs:
                key = prog.upper()
                result.setdefault(campus, {}).setdefault(course, {})
                result[campus][course][key] = result[campus][course].get(key, 0.0) + per_prog
        else:
            key = str(major_raw).strip().upper()
            if not key:
                return
            result.setdefault(campus, {}).setdefault(course, {})
            result[campus][course][key] = result[campus][course].get(key, 0.0) + enroll

    if path.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheet_name = "FA25 + WI26" if "FA25 + WI26" in wb.sheetnames else wb.sheetnames[0]
            ws = wb[sheet_name]
            header_idx: Dict[str, int] = {}
            header_found = False
            for row in ws.iter_rows(values_only=True):
                if not header_found:
                    if any(str(c).strip().lower() == "term" for c in row if c is not None):
                        for i, h in enumerate(row):
                            if h is not None:
                                header_idx[str(h).strip().lower()] = i
                        header_found = True
                    continue
                if not row or row[0] is None:
                    break
                _accumulate(
                    row[header_idx.get("campus", 2)],
                    row[header_idx.get("course", 1)],
                    row[header_idx.get("major", 3)],
                    row[header_idx.get("enrollment", 4)],
                    True,
                )
        except Exception as exc:
            logging.warning("Failed to parse enrollment xlsx %s: %s", path, exc)
            result.clear()
    else:
        with path.open(newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                _accumulate(
                    row.get("campus", ""),
                    row.get("course", ""),
                    row.get("major", ""),
                    row.get("enrollment", 0),
                    False,
                )

    return result


# Canonical display labels for forecast result rows, keyed by the normalized
# campus codes the loaders produce. One copy: result rows, backtest joins, and
# warnings must all emit identical labels or (course, campus) keys stop matching.
CAMPUS_DISPLAY = {"SAVANNAH": "Savannah", "SCADNOW": "SCADnow", "ATLANTA": "Atlanta"}


def _normalize_campus_label(raw_lower: str) -> str:
    """Normalize free-text campus label (already .strip().lower()) to internal constant."""
    if "scadnow" in raw_lower or "online" in raw_lower or raw_lower == "now":
        return "SCADNOW"
    if raw_lower in ("atl", "atlanta"):
        return "ATLANTA"
    return "SAVANNAH"


def _normalize_campus_code(code_upper: str) -> Optional[str]:
    """Normalize CAMPUS column codes from Master Schedule to internal constants.

    Returns None for unrecognized codes so the caller can skip the row.
    """
    _MAP = {"NOW": "SCADNOW", "SAV": "SAVANNAH", "ATL": "ATLANTA"}
    return _MAP.get(code_upper)


def _process_master_row(
    getter,
    by_term: Dict[str, Dict[Tuple[str, str], float]],
    seen_crns: set,
    xwalk: Dict[str, str],
    metric: str,
    term_code: Optional[str],
    has_crn_column: bool,
) -> None:
    """Accumulate one Master Schedule data row into ``by_term``.

    Shared by the CSV and xlsx loaders so filtering and CRN-dedupe semantics
    cannot drift between export formats. ``metric`` must be pre-normalized.
    In a file that carries a CRN column, a row with a blank CRN is a Cognos
    summary/continuation artifact and is skipped; files without a CRN column
    (older exports) are counted row-by-row with no dedupe.
    """
    term_value = str(getter("TERM") or "").strip()
    if term_code and term_value != str(term_code):
        # Skip but don't mark as seen — future term filters may differ.
        return
    crn_key = None
    if has_crn_column:
        crn_raw = str(getter("CRN") or "").strip()
        if not crn_raw:
            return
        # De-duplicate co-taught rows: only count each term+CRN once.
        crn_key = f"{term_value}:{crn_raw}"
        if crn_key in seen_crns:
            return
    subj = str(getter("SUBJ") or "").strip().upper()
    crs = str(getter("CRS NUMBER") or "").strip()
    if not crs:
        # A continuation row with no course number must NOT mark its CRN as
        # seen, or it suppresses the data-carrying row for the same CRN.
        return
    if crn_key:
        seen_crns.add(crn_key)
    raw_course = f"{subj} {crs}"
    course = xwalk.get(raw_course, raw_course)
    if not course.startswith("FOUN "):
        return
    campus = _normalize_campus_code(str(getter("CAMPUS") or "").strip().upper())
    if campus is None:
        return
    term_totals = by_term.setdefault(term_value, defaultdict(float))
    term_totals[(campus, course)] += _metric_enrollment(getter, metric)


def _load_master_schedule_xlsx_by_term(
    path: Path,
    crosswalk: Optional[Dict[str, str]] = None,
    demand_metric: Optional[str] = "actual",
    term_code: Optional[str] = None,
) -> Dict[str, Dict[Tuple[str, str], float]]:
    """Single-pass load of a PZSMSCP xlsx, grouped by term code.

    Format expectations:
      - Single data sheet (typically named ``Page1``)
      - Rows 0-15 are Cognos metadata; the header row contains 'SCHOOL', 'CRN',
        'SUBJ', 'CRS NUMBER', 'TERM', 'CAMPUS', 'ACT ENR' (located by name)
      - Multiple rows per CRN are common (one per instructor when co-taught);
        rows are de-duplicated by term+CRN before aggregation.
    """
    by_term: Dict[str, Dict[Tuple[str, str], float]] = {}
    xwalk = crosswalk or {}
    metric = normalize_demand_metric(demand_metric)

    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
    except Exception as exc:
        logging.warning("Failed to open Master Schedule xlsx %s: %s", path, exc)
        return {}

    # Locate the header row by scanning for a row that contains the canonical
    # column names. Cognos puts ~15 rows of report metadata above the header.
    header_idx: Dict[str, int] = {}
    in_data = False
    has_crn = False
    seen_crns: set = set()

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        if not in_data:
            # A header row contains 'SUBJ' and 'CRS NUMBER' literals
            cells = [str(c).strip().upper() if c is not None else "" for c in row]
            if "SUBJ" in cells and "CRS NUMBER" in cells and "ACT ENR" in cells:
                for i, c in enumerate(cells):
                    if c:
                        header_idx[c] = i
                has_crn = "CRN" in header_idx
                in_data = True
            continue

        def _cell(name: str):
            i = header_idx.get(name)
            if i is None or i >= len(row):
                return None
            return row[i]

        _process_master_row(_cell, by_term, seen_crns, xwalk, metric, term_code, has_crn)

    return by_term


def _load_master_schedule_xlsx(
    path: Path,
    term_code: Optional[str] = None,
    crosswalk: Optional[Dict[str, str]] = None,
    demand_metric: Optional[str] = "actual",
) -> Dict[Tuple[str, str], float]:
    """Load flat enrollment totals from a PZSMSCP Master Schedule xlsx."""
    by_term = _load_master_schedule_xlsx_by_term(
        path, crosswalk=crosswalk, demand_metric=demand_metric, term_code=term_code
    )
    if term_code:
        return by_term.get(str(term_code), {})
    totals: Dict[Tuple[str, str], float] = defaultdict(float)
    for term_totals in by_term.values():
        for key, seats in term_totals.items():
            totals[key] += seats
    return totals


def load_term_enrollments(
    path: Path,
    term_code: Optional[str] = None,
    crosswalk: Optional[Dict[str, str]] = None,
    demand_metric: Optional[str] = "actual",
) -> Dict[Tuple[str, str], float]:
    """Load enrollment totals per (campus, course) from a term CSV, Master Schedule CSV, or Master Schedule xlsx.

    Dispatches by file extension. xlsx is assumed to be the PZSMSCP Cognos
    Master Schedule export (the only xlsx format the tool currently consumes
    for enrollment data).

    Args:
        crosswalk: Optional legacy→FOUN code mapping. When provided, rows with
            legacy subject codes (DRAW, DSGN, etc.) are mapped to their FOUN
            equivalents so feeder enrollment is not silently dropped.
    """
    if path.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
        return _load_master_schedule_xlsx(
            path,
            term_code=term_code,
            crosswalk=crosswalk,
            demand_metric=demand_metric,
        )

    by_term = _load_master_schedule_csv_by_term(
        path, crosswalk=crosswalk, demand_metric=demand_metric, term_code=term_code
    )
    if by_term is not None:
        if term_code:
            return by_term.get(str(term_code), {})
        totals: Dict[Tuple[str, str], float] = defaultdict(float)
        for term_totals in by_term.values():
            for key, seats in term_totals.items():
                totals[key] += seats
        return totals

    # Legacy per-term CSV: 'Course' / 'Enrollment' columns, no TERM column.
    totals = defaultdict(float)
    xwalk = crosswalk or {}
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.DictReader(f)
        fieldnames = [name or "" for name in (reader.fieldnames or [])]
        has_course = "Course" in fieldnames and "Enrollment" in fieldnames
        if not has_course:
            return totals
        for row in reader:
            course = (row.get("Course") or "").strip()
            course = xwalk.get(course, course)
            if not course.startswith("FOUN "):
                continue
            enrollment = parse_number(row.get("Enrollment"))
            room = (row.get("Room") or "").strip().upper()
            section = (row.get("Section #") or "").strip().upper()
            # Also check the explicit Campus column when present (e.g. "SCADnow online")
            campus_col = (row.get("Campus") or "").strip().lower()
            if campus_col:
                campus = _normalize_campus_label(campus_col)
            else:
                campus = "SCADNOW" if (room == "OLNOW" or section.startswith("N")) else "SAVANNAH"
            totals[(campus, course)] += enrollment
    return totals


def _load_master_schedule_csv_by_term(
    path: Path,
    crosswalk: Optional[Dict[str, str]] = None,
    demand_metric: Optional[str] = "actual",
    term_code: Optional[str] = None,
) -> Optional[Dict[str, Dict[Tuple[str, str], float]]]:
    """Single-pass load of a Master Schedule CSV, grouped by term code.

    Returns None when the file is not in Master Schedule format (no SUBJ /
    CRS NUMBER / ACT ENR columns), so the caller can fall back to the legacy
    per-term CSV format.
    """
    xwalk = crosswalk or {}
    metric = normalize_demand_metric(demand_metric)
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.DictReader(f)
        fieldnames = [name or "" for name in (reader.fieldnames or [])]
        has_master = "SUBJ" in fieldnames and "CRS NUMBER" in fieldnames and "ACT ENR" in fieldnames
        if not has_master:
            return None
        has_crn = "CRN" in fieldnames
        by_term: Dict[str, Dict[Tuple[str, str], float]] = {}
        seen_crns: set = set()
        for row in reader:
            _process_master_row(row.get, by_term, seen_crns, xwalk, metric, term_code, has_crn)
        return by_term


def load_all_term_enrollments(
    path: Path,
    crosswalk: Optional[Dict[str, str]] = None,
    demand_metric: Optional[str] = "actual",
) -> Dict[str, Dict[Tuple[str, str], float]]:
    """Load every term's (campus, course) totals from a Master Schedule in ONE pass.

    Use this instead of calling ``load_term_enrollments`` once per term when a
    full per-term history is needed (e.g. anomaly-detection series): each
    per-term call re-parses the whole file. Legacy per-term CSVs carry no TERM
    column and return {}.
    """
    if path.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
        return _load_master_schedule_xlsx_by_term(
            path, crosswalk=crosswalk, demand_metric=demand_metric
        )
    by_term = _load_master_schedule_csv_by_term(
        path, crosswalk=crosswalk, demand_metric=demand_metric
    )
    return by_term if by_term is not None else {}


def compute_sections(seats: float, capacity: int) -> int:
    if seats <= 0:
        return 0
    return int(math.ceil(seats / capacity))


def distribute_enrollments(
    enrollments: Dict[str, float],
    mapping: Dict[Tuple[str, str], float],
    multiplier: float,
    source_totals: Optional[Dict[str, float]] = None,
) -> Dict[str, float]:
    demand: Dict[str, float] = defaultdict(float)
    by_source: Dict[str, Dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for (source, target), weight in mapping.items():
        by_source[source][target] += weight

    for source_course, seats in enrollments.items():
        if seats <= 0:
            continue
        targets = by_source.get(source_course)
        if not targets:
            continue
        # If a source_totals normalizer is provided, use it so that the proportion
        # reflects the share of programs routing through this source, not just
        # the subset captured by this mapping.
        total_weight = (
            source_totals.get(source_course, 0.0)
            if source_totals is not None
            else sum(targets.values())
        )
        if total_weight <= 0:
            continue
        for target_course, weight in targets.items():
            demand[target_course] += seats * multiplier * (weight / total_weight)
    return demand


def get_available_terms(master_schedule_path: Path) -> List[str]:
    """Scan the Master Schedule (CSV or xlsx) for distinct TERM values.

    Dispatches by file extension. xlsx is assumed to be the PZSMSCP Cognos
    Master Schedule export, which has ~15 rows of metadata before the header.
    """
    terms: set = set()

    if master_schedule_path.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(master_schedule_path, read_only=True, data_only=True)
            ws = wb.active
        except Exception as exc:
            logging.warning("Failed to open Master Schedule xlsx %s: %s", master_schedule_path, exc)
            return []
        in_data = False
        term_idx: Optional[int] = None
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            if not in_data:
                cells = [str(c).strip().upper() if c is not None else "" for c in row]
                if "SUBJ" in cells and "CRS NUMBER" in cells and "TERM" in cells:
                    term_idx = cells.index("TERM")
                    in_data = True
                continue
            if term_idx is None or term_idx >= len(row):
                continue
            val = row[term_idx]
            if val is None:
                continue
            s = str(val).strip()
            if s:
                terms.add(s)
        return sorted(terms)

    with master_schedule_path.open(newline="", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.DictReader(f)
        for row in reader:
            term_val = str(row.get("TERM") or "").strip()
            if term_val:
                terms.add(term_val)
    return sorted(terms)


def term_code_to_label(term_code: str) -> str:
    """Convert SCAD term code like '202630' to 'Spring 2026'."""
    if len(term_code) != 6:
        return term_code
    acad_year = int(term_code[:4])
    quarter_digit = term_code[4:]
    labels = {"10": "Fall", "20": "Winter", "30": "Spring", "40": "Summer"}
    quarter_name = labels.get(quarter_digit)
    if not quarter_name:
        return term_code
    # Fall uses academic year - 1 for calendar year
    if quarter_name == "Fall":
        cal_year = acad_year - 1
    else:
        cal_year = acad_year
    return f"{quarter_name} {cal_year}"


# --------------- Orchestrator ---------------

def run_sequence_forecast(
    sequence_map_path: Path,
    enrollment_source_path: Path,
    target_term: str,
    capacity: int = 20,
    progression_rate: float = 0.95,
    buffer_percent: float = 0.0,
    admits_path: Optional[Path] = None,
    enrollment_by_major_path: Optional[Path] = None,
    demand_metric: Optional[str] = "actual",
) -> List[Dict]:
    """Run the full sequence-based forecast pipeline for any target quarter.

    Args:
        sequence_map_path: Path to the sequencing map CSV.
        enrollment_source_path: Path to enrollment data (Master Schedule or term CSV).
        target_term: Human-readable term, e.g. "Summer 2026".
        capacity: Section capacity.
        progression_rate: Per-gap progression rate.
        buffer_percent: Buffer percentage to add.

    Returns a list of dicts with keys:
        course, campus, projected_seats, sections, method
    """
    info = resolve_term_info(target_term)
    target_quarter = info["target_quarter"]
    closer = info["closer_feeder"]
    farther = info["farther_feeder"]

    year_filter = _active_curriculum_years(info["target_term_code"])

    # Load optional enrollment-by-major weights for sequence map
    enrollment_weights = None
    if enrollment_by_major_path:
        enrollment_weights = load_enrollment_by_major(enrollment_by_major_path)

    # Load admits demand once (avoids re-reading file per campus)
    admits_demand = load_admits_foun_demand(admits_path) if admits_path else {}

    mappings = load_sequence_mappings(
        sequence_map_path,
        target_quarter=target_quarter,
        closer_quarter=closer["quarter"],
        farther_quarter=farther["quarter"],
        year_filter=year_filter,
        enrollment_weights=enrollment_weights,
    )

    # Load crosswalk so legacy course codes (DRAW, DSGN) map to FOUN
    crosswalk_path = enrollment_source_path.parent / "sequence_crosswalk_template.csv"
    crosswalk = load_crosswalk(crosswalk_path)

    farther_enrollments = load_term_enrollments(
        enrollment_source_path,
        farther["term_code"],
        crosswalk=crosswalk,
        demand_metric=demand_metric,
    )
    closer_enrollments = load_term_enrollments(
        enrollment_source_path,
        closer["term_code"],
        crosswalk=crosswalk,
        demand_metric=demand_metric,
    )

    farther_multiplier = progression_rate ** farther["multiplier_exp"]
    closer_multiplier = progression_rate ** closer["multiplier_exp"]

    output_rows: List[Dict] = []
    for campus in ("SAVANNAH", "SCADNOW", "ATLANTA"):
        farther_by_course = {
            course: seats
            for (campus_key, course), seats in farther_enrollments.items()
            if campus_key == campus
        }
        closer_by_course = {
            course: seats
            for (campus_key, course), seats in closer_enrollments.items()
            if campus_key == campus
        }

        from_farther = distribute_enrollments(
            farther_by_course,
            mappings[campus]["farther_to_target"],
            farther_multiplier,
            source_totals=mappings[campus]["farther_source_totals"],
        )
        from_closer = distribute_enrollments(
            closer_by_course,
            mappings[campus]["closer_to_target"],
            closer_multiplier,
            source_totals=mappings[campus]["closer_source_totals"],
        )

        combined = defaultdict(float)
        for course, seats in from_farther.items():
            combined[course] += seats
        for course, seats in from_closer.items():
            combined[course] += seats

        # Add new-admit FOUN demand (intro courses entered directly from admits file)
        for course, count in admits_demand.get(campus, {}).items():
            combined[course] += count

        for course in mappings[campus]["target_counts"].keys():
            combined.setdefault(course, 0.0)

        # Apply buffer
        buffer_multiplier = 1.0 + (buffer_percent / 100.0)

        for course in sorted(combined.keys()):
            seats = combined[course] * buffer_multiplier
            output_rows.append(
                {
                    "course": course,
                    "campus": CAMPUS_DISPLAY.get(campus, campus),
                    "projected_seats": seats,
                    "sections": compute_sections(seats, capacity),
                    "method": "sequence_map_feeder_mapping",
                }
            )

    return output_rows


def _post_rollout_prior_same_season_codes(target_term: str) -> List[str]:
    """Post-rollout prior same-season term codes for the historical method.

    Returns the same-quarter term codes (most recent first) whose academic year
    is after _FOUN_CURRICULUM_START. Empty when target_term is the first
    post-rollout instance of its season, meaning no prior same-season term under
    the current curriculum exists to forecast from.
    """
    info = resolve_term_info(target_term)
    target_code = info["target_term_code"]
    yyyy = int(target_code[:4])
    qq = target_code[4:]
    codes: List[str] = []
    k = 1
    while (yyyy - k) > _FOUN_CURRICULUM_START:
        codes.append(f"{yyyy - k}{qq}")
        k += 1
    return codes


def run_sameseason_forecast(
    master_schedule_path: Path,
    target_term: str,
    capacity: int = 20,
    buffer_percent: float = 0.0,
    crosswalk: Optional[Dict[str, str]] = None,
    demand_metric: Optional[str] = "actual",
) -> List[Dict]:
    """Forecast each FOUN course from its own prior same-season enrollment.

    Robust to sequencing-guide changes: it never reads the sequencing map.
    Uses only post-rollout terms (academic-year part of the code
    > _FOUN_CURRICULUM_START). With 2+ prior same-season points it fits the
    season-aware OLS trend; with one point it uses that value (level).
    Assumes the prior same-season terms are consecutive (no gap years), which
    holds for the post-rollout window. Courses offered in the target term but
    with no same-season history are returned with projected_seats=0 and method
    "same_season_new_course" so the admin can set a manual estimate.

    Returns rows shaped like run_sequence_forecast:
        {course, campus, projected_seats, sections, method}
    Returns [] when no same-season history (and no target-term courses) exist;
    the caller turns that into a clear guardrail error.
    """
    import pandas as pd
    from forecast_tool.forecasting.ols_forecast import forecast_ols

    if crosswalk is None:
        crosswalk = load_crosswalk(master_schedule_path.parent / "sequence_crosswalk_template.csv")

    info = resolve_term_info(target_term)
    target_code = info["target_term_code"]

    prior_codes = _post_rollout_prior_same_season_codes(target_term)

    series: DefaultDict[Tuple[str, str], Dict[int, float]] = defaultdict(dict)
    for code in prior_codes:
        for (campus, course), seats in load_term_enrollments(
            master_schedule_path,
            code,
            crosswalk=crosswalk,
            demand_metric=demand_metric,
        ).items():
            if str(course).startswith("FOUN"):
                series[(campus, course)][int(code[:4])] = seats

    campus_label = CAMPUS_DISPLAY
    buffer_mult = 1.0 + (buffer_percent / 100.0)
    rows: List[Dict] = []

    for (campus, course), year_map in series.items():
        pts = sorted(year_map.items())
        if len(pts) >= 2:
            df = pd.DataFrame({"ds": [str(y) for y, _ in pts], "y": [e for _, e in pts]})
            # Extrapolate to the target's academic year, so a missing
            # intermediate prior does not shift the prediction a year short.
            steps = max(1, int(target_code[:4]) - int(pts[-1][0]))
            fc = forecast_ols(df, periods=steps)
            value = float(fc.iloc[-1]["yhat"]) if not fc.empty else float(pts[-1][1])
        else:
            value = float(pts[-1][1])
        seats = value * buffer_mult
        rows.append({
            "course": course,
            "campus": campus_label.get(campus, campus),
            "projected_seats": seats,
            "sections": compute_sections(seats, capacity),
            "method": "same_season",
        })

    for (campus, course), _seats in load_term_enrollments(
        master_schedule_path,
        target_code,
        crosswalk=crosswalk,
        demand_metric=demand_metric,
    ).items():
        if str(course).startswith("FOUN") and (campus, course) not in series:
            rows.append({
                "course": course,
                "campus": campus_label.get(campus, campus),
                "projected_seats": 0.0,
                "sections": 0,
                "method": "same_season_new_course",
            })

    return rows


def _compute_historical_ratios(
    historical_path: Path,
    target_quarter_code: str,
    feeder_quarter_code: str,
    crosswalk_path: Optional[Path] = None,
) -> Dict[str, float]:
    """Compute average target/feeder enrollment ratios per course from historical data.

    Quarter codes: "10"=Fall, "20"=Winter, "30"=Spring, "40"=Summer.
    Returns {course: ratio}.
    """
    # Load legacy→FOUN crosswalk so DRAW/DSGN rows map to FOUN codes
    crosswalk = load_crosswalk(crosswalk_path) if crosswalk_path else {}

    # Collect per-course, per-academic-year enrollment totals for each quarter
    # Structure: {course: {acad_year: {quarter_code: total_enrollment}}}
    data: Dict[str, Dict[str, Dict[str, float]]] = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

    if not historical_path.is_file():
        return {}

    with historical_path.open(newline="", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.DictReader(f)
        for row in reader:
            subj = (row.get("SUBJ") or "").strip().upper()
            crs = (row.get("CRS NUMBER") or "").strip()
            if not crs:
                continue
            raw_course = f"{subj} {crs}"
            # Map legacy codes to FOUN codes; keep FOUN codes as-is
            course = crosswalk.get(raw_course, raw_course)
            if not course.startswith("FOUN "):
                continue
            term_str = str(row.get("TERM") or "").strip()
            if len(term_str) != 6:
                continue
            acad_year = term_str[:4]
            qq = term_str[4:]
            enrollment = parse_number(row.get("ACT ENR"))
            data[course][acad_year][qq] += enrollment

    ratios: Dict[str, float] = {}
    for course, by_year in data.items():
        year_ratios = []
        for acad_year, by_qq in by_year.items():
            feeder_enr = by_qq.get(feeder_quarter_code, 0.0)
            target_enr = by_qq.get(target_quarter_code, 0.0)
            if feeder_enr > 0 and target_enr > 0:
                year_ratios.append(target_enr / feeder_enr)
        if year_ratios:
            ratios[course] = sum(year_ratios) / len(year_ratios)

    return ratios


def run_ratio_forecast(
    feeder_forecast_path: Path,
    historical_data_path: Path,
    target_term: str,
    capacity: int = 20,
    buffer_percent: float = 0.0,
    default_ratio: float = 0.12,
) -> List[Dict]:
    """Ratio-based forecast: apply historical target/feeder ratios to a prior forecast.

    Used when the sequence map lacks data for the target quarter (e.g. Summer).
    Reads an existing forecast CSV (e.g. Spring 2026) and scales seats by
    the historical ratio of target-quarter to feeder-quarter enrollment.

    Args:
        feeder_forecast_path: Path to the feeder term's forecast CSV
            (must have columns: course, campus, and a *_projected_seats column).
        historical_data_path: Path to FOUN_Historical.csv for ratio computation.
        target_term: Human-readable term, e.g. "Summer 2026".
        capacity: Section capacity.
        buffer_percent: Buffer percentage to add.
        default_ratio: Fallback ratio when historical data is insufficient.

    Returns list of dicts matching run_sequence_forecast output format.
    """
    info = resolve_term_info(target_term)
    target_qq = str(QUARTER_CODES[info["target_quarter"]])
    feeder_qq = str(QUARTER_CODES[info["closer_feeder"]["quarter"]])

    # Compute per-course historical ratios (with legacy code crosswalk)
    crosswalk_path = historical_data_path.parent / "sequence_crosswalk_template.csv"
    historical_ratios = _compute_historical_ratios(
        historical_data_path, target_qq, feeder_qq, crosswalk_path=crosswalk_path
    )

    # Load the feeder forecast CSV
    feeder_data: List[Tuple[str, str, float]] = []
    if not feeder_forecast_path.is_file():
        return []

    with feeder_forecast_path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []
        # Find the projected_seats column (varies by CSV naming)
        seats_col = None
        for col in fieldnames:
            if col == "projected_seats" or col.endswith("_projected_seats"):
                seats_col = col
                break
        if seats_col is None:
            return []

        for row in reader:
            course = (row.get("course") or "").strip()
            campus = (row.get("campus") or "").strip()
            seats = parse_number(row.get(seats_col))
            if course and campus and seats > 0:
                feeder_data.append((course, campus, seats))

    buffer_multiplier = 1.0 + (buffer_percent / 100.0)
    output_rows: List[Dict] = []

    for course, campus, feeder_seats in feeder_data:
        ratio = historical_ratios.get(course, default_ratio)
        projected = feeder_seats * ratio * buffer_multiplier
        sections = compute_sections(projected, capacity)
        if sections > 0:
            output_rows.append({
                "course": course,
                "campus": campus,
                "projected_seats": projected,
                "sections": sections,
                "method": "ratio_based",
            })

    return output_rows


def load_previous_forecast(csv_path: Path) -> Dict[Tuple[str, str], float]:
    """Read an existing forecast CSV to compute change deltas.

    Returns {(course, campus): projected_seats}.
    Tries 'projected_seats' column first, falls back to columns ending
    with '_projected_seats' for backward compat with old CSVs.
    """
    result: Dict[Tuple[str, str], float] = {}
    if not csv_path.is_file():
        return result
    with csv_path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []
        # Find the projected_seats column
        seats_col = None
        if "projected_seats" in fieldnames:
            seats_col = "projected_seats"
        else:
            for col in fieldnames:
                if col.endswith("_projected_seats"):
                    seats_col = col
                    break
        if seats_col is None:
            return result
        for row in reader:
            course = (row.get("course") or "").strip()
            campus = (row.get("campus") or "").strip()
            seats = parse_number(row.get(seats_col))
            if course and campus:
                result[(course, campus)] = seats
    return result
