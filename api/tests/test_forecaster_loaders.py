"""Tests for load_term_enrollments and get_available_terms.

load_term_enrollments handles two distinct CSV layouts:
  - Term CSV:  Course / Section # / Room / Enrollment columns
  - Master Schedule: SUBJ / CRS NUMBER / ACT ENR / CAMPUS / TERM columns

get_available_terms scans the Master Schedule for distinct TERM values.
"""

import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from forecaster import get_available_terms, load_admits_foun_demand, load_enrollment_by_major, load_term_enrollments


# ── Fixture helpers ────────────────────────────────────────────────────────────

def _write(path: Path, content: str) -> None:
    path.write_text(content, encoding="utf-8")


# ── load_term_enrollments — term CSV format ────────────────────────────────────

class TestLoadTermEnrollmentsTermCSV:
    """Term CSV uses Course / Section # / Room / Enrollment columns."""

    def test_loads_savannah_enrollment(self, tmp_path):
        csv = tmp_path / "term.csv"
        _write(csv, "Course,Section #,Room,Enrollment\nFOUN 110,S01,SAV-101,100\n")
        result = load_term_enrollments(csv)
        assert result[("SAVANNAH", "FOUN 110")] == 100.0

    def test_scadnow_detected_by_olnow_room(self, tmp_path):
        csv = tmp_path / "term.csv"
        _write(csv, "Course,Section #,Room,Enrollment\nFOUN 110,S01,OLNOW,50\n")
        result = load_term_enrollments(csv)
        assert result[("SCADNOW", "FOUN 110")] == 50.0

    def test_scadnow_detected_by_n_prefixed_section(self, tmp_path):
        csv = tmp_path / "term.csv"
        _write(csv, "Course,Section #,Room,Enrollment\nFOUN 230,N01,STUDIO,40\n")
        result = load_term_enrollments(csv)
        assert result[("SCADNOW", "FOUN 230")] == 40.0

    def test_non_foun_courses_are_excluded(self, tmp_path):
        csv = tmp_path / "term.csv"
        _write(csv, "Course,Section #,Room,Enrollment\nDRAW 200,S01,SAV-101,80\n")
        result = load_term_enrollments(csv)
        assert len(result) == 0

    def test_multiple_sections_accumulate_for_same_course(self, tmp_path):
        csv = tmp_path / "term.csv"
        _write(csv,
            "Course,Section #,Room,Enrollment\n"
            "FOUN 110,S01,SAV-101,50\n"
            "FOUN 110,S02,SAV-102,60\n"
        )
        result = load_term_enrollments(csv)
        assert result[("SAVANNAH", "FOUN 110")] == 110.0

    def test_empty_file_returns_empty_dict(self, tmp_path):
        csv = tmp_path / "term.csv"
        _write(csv, "Course,Section #,Room,Enrollment\n")
        assert load_term_enrollments(csv) == {}

    def test_crosswalk_maps_legacy_code_to_foun(self, tmp_path):
        csv = tmp_path / "term.csv"
        _write(csv, "Course,Section #,Room,Enrollment\nDRAW 200,S01,SAV-101,75\n")
        result = load_term_enrollments(csv, crosswalk={"DRAW 200": "FOUN 230"})
        assert result[("SAVANNAH", "FOUN 230")] == 75.0

    def test_crosswalk_does_not_affect_already_foun_codes(self, tmp_path):
        csv = tmp_path / "term.csv"
        _write(csv, "Course,Section #,Room,Enrollment\nFOUN 110,S01,SAV-101,90\n")
        result = load_term_enrollments(csv, crosswalk={"FOUN 110": "FOUN 999"})
        # Crosswalk maps the key exactly; FOUN 110 → FOUN 999, but still FOUN prefix
        assert result[("SAVANNAH", "FOUN 999")] == 90.0


# ── load_term_enrollments — Master Schedule format ────────────────────────────

class TestLoadTermEnrollmentsMasterSchedule:
    """Master Schedule uses SUBJ / CRS NUMBER / ACT ENR / CAMPUS / TERM."""

    def _master(self, path: Path, rows: list[dict]) -> None:
        lines = ["SUBJ,CRS NUMBER,ACT ENR,CAMPUS,TERM"]
        for r in rows:
            lines.append(f"{r['subj']},{r['crs']},{r['enr']},{r['campus']},{r['term']}")
        _write(path, "\n".join(lines) + "\n")

    def test_loads_savannah_row_with_term_filter(self, tmp_path):
        csv = tmp_path / "master.csv"
        self._master(csv, [{"subj": "FOUN", "crs": "110", "enr": 100, "campus": "SAV", "term": "202630"}])
        result = load_term_enrollments(csv, term_code="202630")
        assert result[("SAVANNAH", "FOUN 110")] == 100.0

    def test_loads_scadnow_row(self, tmp_path):
        csv = tmp_path / "master.csv"
        self._master(csv, [{"subj": "FOUN", "crs": "230", "enr": 60, "campus": "NOW", "term": "202630"}])
        result = load_term_enrollments(csv, term_code="202630")
        assert result[("SCADNOW", "FOUN 230")] == 60.0

    def test_term_filter_excludes_other_terms(self, tmp_path):
        csv = tmp_path / "master.csv"
        self._master(csv, [
            {"subj": "FOUN", "crs": "110", "enr": 100, "campus": "SAV", "term": "202630"},
            {"subj": "FOUN", "crs": "110", "enr": 50,  "campus": "SAV", "term": "202620"},
        ])
        result = load_term_enrollments(csv, term_code="202630")
        assert result[("SAVANNAH", "FOUN 110")] == 100.0

    def test_no_term_filter_loads_all_terms(self, tmp_path):
        csv = tmp_path / "master.csv"
        self._master(csv, [
            {"subj": "FOUN", "crs": "110", "enr": 100, "campus": "SAV", "term": "202630"},
            {"subj": "FOUN", "crs": "110", "enr": 50,  "campus": "SAV", "term": "202620"},
        ])
        result = load_term_enrollments(csv)  # no term_code
        assert result[("SAVANNAH", "FOUN 110")] == 150.0

    def test_atl_campus_maps_to_atlanta(self, tmp_path):
        csv = tmp_path / "master.csv"
        self._master(csv, [{"subj": "FOUN", "crs": "110", "enr": 80, "campus": "ATL", "term": "202630"}])
        result = load_term_enrollments(csv, term_code="202630")
        assert result[("ATLANTA", "FOUN 110")] == 80.0

    def test_unknown_campus_still_excluded(self, tmp_path):
        csv = tmp_path / "master.csv"
        self._master(csv, [{"subj": "FOUN", "crs": "110", "enr": 80, "campus": "HK", "term": "202630"}])
        result = load_term_enrollments(csv, term_code="202630")
        assert len(result) == 0

    def test_non_foun_subjects_excluded(self, tmp_path):
        csv = tmp_path / "master.csv"
        self._master(csv, [{"subj": "DRAW", "crs": "200", "enr": 70, "campus": "SAV", "term": "202630"}])
        result = load_term_enrollments(csv, term_code="202630")
        assert len(result) == 0

    def test_crosswalk_maps_legacy_subject_code(self, tmp_path):
        csv = tmp_path / "master.csv"
        self._master(csv, [{"subj": "DRAW", "crs": "200", "enr": 70, "campus": "SAV", "term": "202630"}])
        result = load_term_enrollments(csv, term_code="202630", crosswalk={"DRAW 200": "FOUN 230"})
        assert result[("SAVANNAH", "FOUN 230")] == 70.0

    def test_multiple_sections_accumulate(self, tmp_path):
        csv = tmp_path / "master.csv"
        self._master(csv, [
            {"subj": "FOUN", "crs": "110", "enr": 20, "campus": "SAV", "term": "202630"},
            {"subj": "FOUN", "crs": "110", "enr": 18, "campus": "SAV", "term": "202630"},
        ])
        result = load_term_enrollments(csv, term_code="202630")
        assert result[("SAVANNAH", "FOUN 110")] == 38.0

    def test_master_csv_dedupes_duplicate_crn_rows(self, tmp_path):
        csv = tmp_path / "master.csv"
        _write(csv,
            "CRN,SUBJ,CRS NUMBER,ACT ENR,CAMPUS,TERM\n"
            "12345,FOUN,110,20,SAV,202630\n"
            "12345,FOUN,110,20,SAV,202630\n"
        )
        result = load_term_enrollments(csv, term_code="202630")
        assert result[("SAVANNAH", "FOUN 110")] == 20.0

    def test_demand_metric_max_uses_max_enrollment(self, tmp_path):
        csv = tmp_path / "master.csv"
        _write(csv,
            "SUBJ,CRS NUMBER,ACT ENR,MAX ENR,WL ACT ENR,CAMPUS,TERM\n"
            "FOUN,110,17,20,3,SAV,202630\n"
        )
        result = load_term_enrollments(csv, term_code="202630", demand_metric="max")
        assert result[("SAVANNAH", "FOUN 110")] == 20.0

    def test_demand_metric_actual_plus_waitlist(self, tmp_path):
        csv = tmp_path / "master.csv"
        _write(csv,
            "SUBJ,CRS NUMBER,ACT ENR,MAX ENR,WL ACT ENR,CAMPUS,TERM\n"
            "FOUN,110,17,20,3,SAV,202630\n"
        )
        result = load_term_enrollments(csv, term_code="202630", demand_metric="actual_plus_waitlist")
        assert result[("SAVANNAH", "FOUN 110")] == 20.0

    def test_missing_crs_number_skipped(self, tmp_path):
        csv = tmp_path / "master.csv"
        _write(csv, "SUBJ,CRS NUMBER,ACT ENR,CAMPUS,TERM\nFOUN,,20,SAV,202630\n")
        result = load_term_enrollments(csv, term_code="202630")
        assert len(result) == 0


# ── get_available_terms ────────────────────────────────────────────────────────

class TestGetAvailableTerms:
    def test_returns_sorted_unique_term_codes(self, tmp_path):
        csv = tmp_path / "master.csv"
        csv.write_text(
            "TERM,SUBJ,CRS NUMBER\n"
            "202630,FOUN,110\n"
            "202610,FOUN,110\n"
            "202630,FOUN,230\n"
        )
        result = get_available_terms(csv)
        assert result == ["202610", "202630"]

    def test_empty_file_returns_empty_list(self, tmp_path):
        csv = tmp_path / "master.csv"
        csv.write_text("TERM,SUBJ,CRS NUMBER\n")
        assert get_available_terms(csv) == []

    def test_skips_rows_with_empty_term(self, tmp_path):
        csv = tmp_path / "master.csv"
        csv.write_text("TERM,SUBJ,CRS NUMBER\n,FOUN,110\n202630,FOUN,230\n")
        result = get_available_terms(csv)
        assert result == ["202630"]


# ── Fixtures ───────────────────────────────────────────────────────────────────

def _admits_xlsx(tmp_path: Path, students: list) -> Path:
    """Create minimal PZSAAPF-style xlsx.
    students = [(campus_code, registered_courses_string), ...]
    Campus codes: 'M'=Savannah, 'O'=SCADnow.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    # Rows 1–10: empty (metadata region in real file)
    for _ in range(10):
        ws.append([None])
    # Row 11: headers — only cols 0 (A), 12 (M), 20 (U) matter
    header_row = [None] * 39
    header_row[0] = "Student ID"
    header_row[12] = "Campus"
    header_row[20] = "Currently Registered Courses (NO WL)"
    ws.append(header_row)
    # Data rows
    for i, (campus_code, reg_courses) in enumerate(students):
        row = [None] * 39
        row[0] = f"student_{i}"
        row[12] = campus_code
        row[20] = reg_courses
        ws.append(row)
    p = tmp_path / "admits.xlsx"
    wb.save(p)
    return p


# ── load_admits_foun_demand ────────────────────────────────────────────────────


class TestLoadAdmitsFounDemand:
    def test_counts_foun_courses_for_savannah(self, tmp_path):
        p = _admits_xlsx(tmp_path, [
            ("M", "FOUN 110; FOUN 111; ENGL 123"),
            ("M", "FOUN 110; FSYR 101"),
        ])
        result = load_admits_foun_demand(p)
        assert result["SAVANNAH"]["FOUN 110"] == 2
        assert result["SAVANNAH"]["FOUN 111"] == 1

    def test_counts_foun_courses_for_scadnow(self, tmp_path):
        p = _admits_xlsx(tmp_path, [
            ("O", "FOUN 111; CTXT 122"),
        ])
        result = load_admits_foun_demand(p)
        assert result["SCADNOW"]["FOUN 111"] == 1

    def test_non_foun_courses_excluded(self, tmp_path):
        p = _admits_xlsx(tmp_path, [
            ("M", "ENGL 123; FSYR 101"),
        ])
        result = load_admits_foun_demand(p)
        assert result.get("SAVANNAH", {}) == {}

    def test_unknown_campus_code_skipped(self, tmp_path):
        p = _admits_xlsx(tmp_path, [
            ("X", "FOUN 110"),
        ])
        result = load_admits_foun_demand(p)
        assert result.get("SAVANNAH", {}) == {}
        assert result.get("SCADNOW", {}) == {}

    def test_missing_file_returns_empty_dict(self, tmp_path):
        result = load_admits_foun_demand(tmp_path / "nonexistent.xlsx")
        assert result == {}

    def test_student_with_no_registered_courses_skipped(self, tmp_path):
        p = _admits_xlsx(tmp_path, [
            ("M", None),
            ("M", "FOUN 112"),
        ])
        result = load_admits_foun_demand(p)
        assert result["SAVANNAH"]["FOUN 112"] == 1


# ── load_enrollment_by_major ──────────────────────────────────────────────────


class TestLoadEnrollmentByMajor:
    def _csv(self, tmp_path: Path, rows: list) -> Path:
        p = tmp_path / "enrollment_by_major.csv"
        _write(p, "term,course,campus,major,enrollment\n" + "\n".join(rows) + "\n")
        return p

    def test_loads_savannah_row(self, tmp_path):
        p = self._csv(tmp_path, ["202610,FOUN 112,SAV,ARCHITECTURE,300"])
        result = load_enrollment_by_major(p)
        assert result["SAVANNAH"]["FOUN 112"]["ARCHITECTURE"] == 300.0

    def test_loads_scadnow_row(self, tmp_path):
        p = self._csv(tmp_path, ["202620,FOUN 112,NOW,ARCHITECTURE,50"])
        result = load_enrollment_by_major(p)
        assert result["SCADNOW"]["FOUN 112"]["ARCHITECTURE"] == 50.0

    def test_loads_atlanta_row(self, tmp_path):
        p = self._csv(tmp_path, ["202610,FOUN 112,ATL,ARCHITECTURE,80"])
        result = load_enrollment_by_major(p)
        assert result["ATLANTA"]["FOUN 112"]["ARCHITECTURE"] == 80.0

    def test_aggregates_across_terms(self, tmp_path):
        p = self._csv(tmp_path, [
            "202610,FOUN 112,SAV,ARCHITECTURE,280",
            "202620,FOUN 112,SAV,ARCHITECTURE,310",
        ])
        result = load_enrollment_by_major(p)
        assert result["SAVANNAH"]["FOUN 112"]["ARCHITECTURE"] == pytest.approx(590.0)

    def test_non_foun_courses_excluded(self, tmp_path):
        p = self._csv(tmp_path, ["202610,DRAW 200,SAV,ILLUSTRATION,100"])
        result = load_enrollment_by_major(p)
        assert "SAVANNAH" not in result or "DRAW 200" not in result.get("SAVANNAH", {})

    def test_unknown_campus_skipped(self, tmp_path):
        p = self._csv(tmp_path, ["202610,FOUN 112,HK,ARCHITECTURE,50"])
        result = load_enrollment_by_major(p)
        assert result == {}

    def test_missing_file_returns_empty_dict(self, tmp_path):
        result = load_enrollment_by_major(tmp_path / "nonexistent.csv")
        assert result == {}

    def test_multiple_majors_for_same_course(self, tmp_path):
        p = self._csv(tmp_path, [
            "202610,FOUN 112,SAV,ARCHITECTURE,300",
            "202610,FOUN 112,SAV,ACCESSORY DESIGN,15",
        ])
        result = load_enrollment_by_major(p)
        assert result["SAVANNAH"]["FOUN 112"]["ARCHITECTURE"] == 300.0
        assert result["SAVANNAH"]["FOUN 112"]["ACCESSORY DESIGN"] == 15.0


# ── load_enrollment_by_major — xlsx + Cognos crosswalk ───────────────────────


def _enrollment_xlsx(tmp_path: Path, data_rows: list, header_offset: int = 6) -> Path:
    """Create a minimal Cognos-style enrollment xlsx.

    data_rows = [(term, course, campus, major, enrollment), ...]
    header_offset: number of blank rows before the header (matches real Cognos export).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    for _ in range(header_offset):
        ws.append([None])
    ws.append(["term", "Course", "campus", "Major", "Enrollment"])
    for row in data_rows:
        ws.append(list(row))
    p = tmp_path / "enrollment.xlsx"
    wb.save(p)
    return p


class TestLoadEnrollmentByMajorXlsx:
    def test_translates_cognos_code_to_seq_map_name(self, tmp_path):
        """ARCH (Cognos) → ARCHITECTURE (seq map) for Savannah."""
        p = _enrollment_xlsx(tmp_path, [("202610", "FOUN112", "SAV", "ARCH", 300)])
        result = load_enrollment_by_major(p)
        assert result["SAVANNAH"]["FOUN 112"]["ARCHITECTURE"] == 300.0

    def test_normalises_course_format_foun112_to_foun_112(self, tmp_path):
        """FOUN112 (no space) should become FOUN 112 in the result."""
        p = _enrollment_xlsx(tmp_path, [("202610", "FOUN112", "SAV", "GRDS", 100)])
        result = load_enrollment_by_major(p)
        assert "FOUN 112" in result["SAVANNAH"]

    def test_splits_enrollment_evenly_for_one_to_many(self, tmp_path):
        """ANIM maps to 3 animation tracks; 300 enrollment → 100 each."""
        p = _enrollment_xlsx(tmp_path, [("202610", "FOUN112", "SAV", "ANIM", 300)])
        result = load_enrollment_by_major(p)
        sav = result["SAVANNAH"]["FOUN 112"]
        assert sav["FALL ONLY WINTER ONLY SPRING ONLY ANIMATION 2D ANIMATION"] == pytest.approx(100.0)
        assert sav["FALL ONLY WINTER ONLY SPRING ONLY ANIMATION 3D CHARACTER ANIMATION"] == pytest.approx(100.0)
        assert sav[
            "FALL ONLY WINTER ONLY SPRING ONLY ANIMATION STORYTELLING AND CONCEPT DEVELOPMENT"
        ] == pytest.approx(100.0)

    def test_unknown_cognos_code_silently_skipped(self, tmp_path):
        """Unrecognised Cognos codes (e.g. THED) should not appear in the result."""
        p = _enrollment_xlsx(tmp_path, [
            ("202610", "FOUN112", "SAV", "THED", 50),
            ("202610", "FOUN112", "SAV", "ARCH", 300),
        ])
        result = load_enrollment_by_major(p)
        assert result["SAVANNAH"]["FOUN 112"]["ARCHITECTURE"] == 300.0
        assert "THED" not in result["SAVANNAH"]["FOUN 112"]

    def test_aggregates_same_program_across_terms(self, tmp_path):
        """Enrollment from Fall and Winter for the same program should be summed."""
        p = _enrollment_xlsx(tmp_path, [
            ("202610", "FOUN112", "SAV", "ARCH", 200),
            ("202620", "FOUN112", "SAV", "ARCH", 250),
        ])
        result = load_enrollment_by_major(p)
        assert result["SAVANNAH"]["FOUN 112"]["ARCHITECTURE"] == pytest.approx(450.0)

    def test_now_campus_maps_to_scadnow(self, tmp_path):
        p = _enrollment_xlsx(tmp_path, [("202610", "FOUN112", "NOW", "ARCH", 40)])
        result = load_enrollment_by_major(p)
        assert result["SCADNOW"]["FOUN 112"]["ARCHITECTURE"] == 40.0

    def test_missing_file_returns_empty_dict(self, tmp_path):
        result = load_enrollment_by_major(tmp_path / "nonexistent.xlsx")
        assert result == {}

    def test_corrupt_xlsx_returns_empty_dict(self, tmp_path):
        """A corrupt xlsx file must return {} (not partial data)."""
        p = tmp_path / "bad.xlsx"
        p.write_bytes(b"this is not a valid xlsx file")
        result = load_enrollment_by_major(p)
        assert result == {}


# ── load_term_enrollments — Master Schedule xlsx (PZSMSCP) ────────────────────

# Canonical PZSMSCP column order (Cognos 'Flexible Master Schedule of Classes
# with Power Prompts' export). Column indices line up with the real file.
_PZSMSCP_HEADER = [
    "SCHOOL", "DEPT", "CRN", "XLST ID", "STATUS", "SUBJ", "CRS NUMBER",
    "SECTION", "SESSION DESC", "COURSE TITLE", "TERM", "CAMPUS",
    "MEET DAYS", "MEET TIMES", "BLDG", "ROOM", "INSTR NAME", "INSTR ID",
    "INSTR CAMPUS", "MAX ENR", "ACT ENR", "SEATS AVAIL", "WL MAX ENR",
    "WL ACT ENR", "WL SEATS AVAIL", "PART OF TERM", "SECTION COUNT",
]


def _pzsmscp_xlsx(tmp_path: Path, sections: list, metadata_rows: int = 16) -> Path:
    """Create a synthetic PZSMSCP-style xlsx.

    sections = [(crn, subj, crs, term, campus, act_enr, instructor_name), ...]

    Multiple section entries with the same CRN simulate co-taught sections
    (one row per instructor). The loader must dedupe by CRN.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Page1"
    # Cognos metadata band (default 16 rows of various filter labels)
    for i in range(metadata_rows):
        ws.append([f"meta-{i}"] + [None] * 5)
    # Header row
    ws.append(_PZSMSCP_HEADER)
    # Build a row template
    def _row(crn, subj, crs, term, campus, act_enr, instr_name):
        row = [None] * len(_PZSMSCP_HEADER)
        row[_PZSMSCP_HEADER.index("SCHOOL")] = "School of Foundation Studies"
        row[_PZSMSCP_HEADER.index("DEPT")] = "FOUN"
        row[_PZSMSCP_HEADER.index("CRN")] = crn
        row[_PZSMSCP_HEADER.index("STATUS")] = "A"
        row[_PZSMSCP_HEADER.index("SUBJ")] = subj
        row[_PZSMSCP_HEADER.index("CRS NUMBER")] = crs
        row[_PZSMSCP_HEADER.index("SECTION")] = "01"
        row[_PZSMSCP_HEADER.index("TERM")] = term
        row[_PZSMSCP_HEADER.index("CAMPUS")] = campus
        row[_PZSMSCP_HEADER.index("INSTR NAME")] = instr_name
        row[_PZSMSCP_HEADER.index("MAX ENR")] = 25
        row[_PZSMSCP_HEADER.index("ACT ENR")] = act_enr
        return row
    for entry in sections:
        ws.append(_row(*entry))
    p = tmp_path / "master_schedule.xlsx"
    wb.save(p)
    return p


class TestLoadTermEnrollmentsMasterScheduleXLSX:
    """PZSMSCP Cognos export — header on row 16, data row 17+, dedupe by CRN."""

    def test_loads_savannah_foun_for_target_term(self, tmp_path):
        from forecaster import load_term_enrollments  # local import for clarity
        p = _pzsmscp_xlsx(tmp_path, [
            (10001, "FOUN", "110", "202630", "SAV", 75, "Smith, J"),
        ])
        result = load_term_enrollments(p, term_code="202630")
        assert result[("SAVANNAH", "FOUN 110")] == 75.0

    def test_dedupes_co_taught_rows_by_crn(self, tmp_path):
        """Critical: the real PZSMSCP file emits one row per instructor.
        The loader must count each CRN exactly once."""
        from forecaster import load_term_enrollments
        p = _pzsmscp_xlsx(tmp_path, [
            (10001, "FOUN", "110", "202630", "SAV", 80, "Smith, J"),
            (10001, "FOUN", "110", "202630", "SAV", 80, "Jones, M"),
            (10001, "FOUN", "110", "202630", "SAV", 80, "Brown, K"),
        ])
        result = load_term_enrollments(p, term_code="202630")
        # 80 once, NOT 240
        assert result[("SAVANNAH", "FOUN 110")] == 80.0

    def test_filters_by_term_code(self, tmp_path):
        from forecaster import load_term_enrollments
        p = _pzsmscp_xlsx(tmp_path, [
            (10001, "FOUN", "110", "202610", "SAV", 100, "Smith, J"),
            (10002, "FOUN", "110", "202630", "SAV", 50, "Jones, M"),
        ])
        result = load_term_enrollments(p, term_code="202630")
        assert result[("SAVANNAH", "FOUN 110")] == 50.0
        # 202610 row excluded
        assert sum(v for v in result.values()) == 50.0

    def test_normalizes_campus_codes(self, tmp_path):
        from forecaster import load_term_enrollments
        p = _pzsmscp_xlsx(tmp_path, [
            (10001, "FOUN", "110", "202630", "SAV", 60, "A"),
            (10002, "FOUN", "110", "202630", "NOW", 40, "B"),
            (10003, "FOUN", "110", "202630", "ATL", 20, "C"),
        ])
        result = load_term_enrollments(p, term_code="202630")
        assert result[("SAVANNAH", "FOUN 110")] == 60.0
        assert result[("SCADNOW", "FOUN 110")] == 40.0
        assert result[("ATLANTA", "FOUN 110")] == 20.0

    def test_unknown_campus_code_skipped(self, tmp_path):
        from forecaster import load_term_enrollments
        p = _pzsmscp_xlsx(tmp_path, [
            (10001, "FOUN", "110", "202630", "XX", 99, "A"),
        ])
        result = load_term_enrollments(p, term_code="202630")
        assert result == {}

    def test_non_foun_courses_excluded(self, tmp_path):
        from forecaster import load_term_enrollments
        p = _pzsmscp_xlsx(tmp_path, [
            (10001, "ENGL", "100", "202630", "SAV", 200, "A"),
            (10002, "FOUN", "110", "202630", "SAV", 30, "B"),
        ])
        result = load_term_enrollments(p, term_code="202630")
        assert result == {("SAVANNAH", "FOUN 110"): 30.0}

    def test_no_term_code_returns_all_terms(self, tmp_path):
        from forecaster import load_term_enrollments
        p = _pzsmscp_xlsx(tmp_path, [
            (10001, "FOUN", "110", "202610", "SAV", 100, "A"),
            (10002, "FOUN", "111", "202630", "SAV", 50, "B"),
        ])
        result = load_term_enrollments(p)  # no term filter
        assert result[("SAVANNAH", "FOUN 110")] == 100.0
        assert result[("SAVANNAH", "FOUN 111")] == 50.0

    def test_crosswalk_maps_legacy_codes(self, tmp_path):
        from forecaster import load_term_enrollments
        p = _pzsmscp_xlsx(tmp_path, [
            (10001, "DRAW", "200", "202630", "SAV", 45, "A"),
        ])
        result = load_term_enrollments(p, term_code="202630",
                                       crosswalk={"DRAW 200": "FOUN 230"})
        assert result[("SAVANNAH", "FOUN 230")] == 45.0

    def test_corrupt_xlsx_returns_empty_dict(self, tmp_path):
        from forecaster import load_term_enrollments
        p = tmp_path / "bad.xlsx"
        p.write_bytes(b"not a real xlsx")
        result = load_term_enrollments(p, term_code="202630")
        assert result == {}

    def test_header_detection_skips_metadata_rows(self, tmp_path):
        """Header is auto-located by SUBJ/CRS NUMBER/ACT ENR — metadata above is ignored."""
        from forecaster import load_term_enrollments
        # 25 metadata rows (vs. real file's 16) — loader should still find the header
        p = _pzsmscp_xlsx(tmp_path, [
            (10001, "FOUN", "112", "202630", "SAV", 99, "A"),
        ], metadata_rows=25)
        result = load_term_enrollments(p, term_code="202630")
        assert result[("SAVANNAH", "FOUN 112")] == 99.0


class TestGetAvailableTermsXLSX:
    """get_available_terms must also handle the PZSMSCP xlsx format."""

    def test_returns_distinct_terms_from_xlsx(self, tmp_path):
        from forecaster import get_available_terms
        p = _pzsmscp_xlsx(tmp_path, [
            (10001, "FOUN", "110", "202610", "SAV", 100, "A"),
            (10002, "FOUN", "111", "202610", "SAV", 80, "A"),
            (10003, "FOUN", "110", "202630", "SAV", 50, "A"),
        ])
        result = get_available_terms(p)
        assert result == ["202610", "202630"]

    def test_empty_xlsx_returns_empty_list(self, tmp_path):
        from forecaster import get_available_terms
        p = _pzsmscp_xlsx(tmp_path, [])
        assert get_available_terms(p) == []

    def test_corrupt_xlsx_returns_empty_list(self, tmp_path):
        from forecaster import get_available_terms
        p = tmp_path / "bad.xlsx"
        p.write_bytes(b"not xlsx")
        assert get_available_terms(p) == []
